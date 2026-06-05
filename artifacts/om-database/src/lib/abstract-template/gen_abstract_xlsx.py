import openpyxl, json, sys, re, datetime
from copy import copy
from openpyxl.utils import get_column_letter

TEMPLATE = "/tmp/slim_marsh.xlsx"
WRAP = 108
# Marshalls fingerprints that must NEVER appear in a generated tab:
FINGERPRINTS = ["cochituate","framingham","tjx companies","mega marshall","cadence capital",
  "sec. 4.2","sec. 5.1","sec. 6.1","sec. 7.3","sec. 7.4","sec. 18.8","sec. 13.6",
  "sch b(4","sch b(2","sch b(5","sch b, 4","omnibus amd","demised premises",
  "play it again sports","huntington learning","massage envy"]

def wrap_lines(text, width=WRAP):
    text = re.sub(r"\s+"," ",str(text)).strip()
    out, line = [], ""
    for w in text.split(" "):
        if len(line)+len(w)+1 > width and line: out.append(line); line=w
        else: line=(line+" "+w).strip()
    if line: out.append(line)
    return out or [""]

def fmtdate(v):
    if not v or str(v).strip() in ("Not Provided","Not provided",""): return ""
    try: return datetime.datetime.strptime(str(v)[:10],"%Y-%m-%d")
    except Exception: return v

def gen(abs_path, out_path):
    wb = openpyxl.load_workbook(TEMPLATE); ws = wb["Marshalls(r)"]
    a = json.load(open(abs_path))
    ws.title = (a["tenantName"]+("(r)" if a.get("dba") else ""))[:31]

    # unmerge the lease-notes region so we can rewrite it freely
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= 142: ws.unmerge_cells(str(mr))

    def anchor(addr):
        for mr in ws.merged_cells.ranges:
            if addr in mr: return mr.coord.split(":")[0]
        return addr
    def S(addr,val): ws[anchor(addr)] = val
    def clear(rng):
        for row in ws[rng]:
            for c in row:
                try: c.value=None
                except Exception: pass

    # ---- COMPREHENSIVE CLEAR of every data region (A..R) + citation columns ----
    clear("C2:R6")                      # header values (labels are in A/E/H)
    clear("B10:R18")                    # amendments
    clear("C30:C39"); clear("I30:R39")  # dates + deposit values
    clear("C40:R60")                    # all 3 notice blocks (keep A labels)
    clear("I61:R61")                    # option-table section citations (I61,M61)
    clear("A67:R74")                    # options
    clear("I76:R79")                    # billing header citations (M79)
    clear("A80:R94")                    # billing + deferred/payback + M94
    for ad in ["G95","G110","F97","C97","E97","G97","I97","C107","D107","G107","I103"]:
        S(ad, None)                     # percentage-rent Marshalls data (keep A-col labels)
    clear("A131:R137")                  # breakpoints
    clear("M1:R141")                    # catch-all: citation cols in the entire top half
    clear("A142:R464")                  # lease notes + CAM/RETX/INS grids + bottom matrices (full rebuild)
    S("M3", None)

    # ---- HEADER (mirrors C20/C22/I20/I22/B24 are =+ formulas, auto-fill) ----
    S("C2",a["tenantName"]); S("I2",a.get("dba","")); S("C4",a["center"])
    S("G4",a.get("suite","")); S("I4",a.get("mriNumber","-")); S("C6",a["premisesGLA"])

    # ---- AMENDMENTS ----
    rows=[10,11,12,14,16]
    for i,d in enumerate(a.get("documents",[])[:5]):
        S(f"B{rows[i]}", f'{d["name"]} dated {d.get("date","")}'+(f' — {d.get("description","")}' if d.get("description") else ""))

    # ---- LEASE DATES ----
    dt=a["dates"]
    S("C30",fmtdate(dt["leaseDate"]["value"])); S("C31",dt.get("openDate",{}).get("value","") or "")
    S("C32",fmtdate(dt["leaseCommencement"]["value"])); S("C33",dt.get("cancelDate",{}).get("value","") or "")
    S("C35",fmtdate(a["expiration"])); S("C36",fmtdate(dt["rentStartDate"]["value"]))
    sd=a.get("securityDeposit",{}); S("I35", sd.get("value","None") if isinstance(sd,dict) else "None")

    # ---- NOTICE (Legal block) ----
    na=(a.get("noticeAddresses") or [{}])[0]
    S("C40","Legal"); S("C41",na.get("name","")); S("C42",na.get("attn","")); S("C43",na.get("address1",""))
    S("C45",na.get("city","")); S("G45",na.get("state","")); S("I45",na.get("zip",""))

    # ---- OPTIONS (<=4) ----
    orows=[67,68,69,70]
    for i,o in enumerate(a.get("options",[])[:4]):
        r=orows[i]; S(f"A{r}",o.get("number")); S(f"B{r}",fmtdate(o.get("windowStart","")))
        S(f"C{r}",o.get("optionType","REN")); S(f"D{r}",fmtdate(o.get("noticeDate","")) or "Automatic")
        S(f"E{r}","=+G4"); S(f"F{r}","=+C6")
    if a.get("optionNotes"):
        S("B72","Option Notes:")
        for i,ln in enumerate(wrap_lines(a["optionNotes"])[:6]): S(f"D{72+i}",ln)

    # ---- BILLING (E=annual; D,G are formulas) ----
    rs=a.get("rentSchedule",[]); brows=[80,81,84,86,88,90]; lbl={2:"Option 1",3:"Option 2",4:"Option 3",5:"Option 4"}
    for i,st in enumerate(rs[:6]):
        r=brows[i]
        if i>=2: S(f"A{r-1}", lbl.get(i,f"Option {i-1}"))
        S(f"A{r}",st.get("incomeCategory","RNT")); S(f"B{r}",fmtdate(st.get("periodStart","")))
        S(f"C{r}",fmtdate(st.get("periodEnd",""))); S(f"E{r}",st.get("annualRent",""))
        S(f"D{r}",f"=E{r}/C6"); S(f"G{r}",f"=E{r}/12")

    # ---- PERCENTAGE RENT ----
    pr=a.get("percentageRent",{}); prval=(pr.get("value","") if isinstance(pr,dict) else str(pr))
    none_pct = prval.strip().lower().startswith("none")
    S("A105", "Use Natural Breakpoint: "+("Yes" if pr.get("naturalBreakpoint") else "No"))
    if (not none_pct) and pr.get("naturalBreakpoint"):
        for i,st in enumerate(rs[:6]):
            r=131+i; S(f"A{r}",fmtdate(st.get("periodStart",""))); S(f"B{r}",0.02); S(f"C{r}",f"=E{brows[i]}/B{r}")

    # ---- LEASE NOTES (rebuild from r142, firm's single-line overflow) ----
    refB,refC,refE,refA=ws["B147"],ws["C147"],ws["E147"],ws["A147"]
    cat={"SECDEP":"(Security Deposit)","USE":"(Use Clause)","RADIUS":"(Radius)","LL RADIUS":"(Radius)",
      "EXCLUSI":"(Exclusive Use)","COTENCY":"(Co-tenancy)","KICKLL":"(Kickout/Termination)","KICKTN":"(Kickout/Termination)",
      "NBNC":"(LL Restrictions)","TRSTR":"(Tenant Restrictions)","CAM":"(CAM Method)","RETX":"(Tax Method)","INS":"(Insurance Method)",
      "ASSNSUB":"(Assignment & Sublet)","ALTERTN":"(Alteration Notes)","OPC":"(Operating Covenant)","GO DARK":"(Go Dark Option)",
      "LATECHG":"(Late Fee Notes)","DEFAULT":"(Default)","GUARANT":"(Guaranty)","OTPRCL":"(Outparcel Restrictions)",
      "PURCHSE":"(Purchase Option)","DUES":"(Merchant's Association)","HOLDOVR":"(Holdover)","PYLON":"(Pylon Sign)",
      "ESTOPPEL":"(Estoppel)","SUBORD":"(Subordination)","RELOCAT":"(Relocation)","UTILITY":"(Utilities)","HVAC":"(HVAC)"}
    def cp(dst,ref): dst._style=copy(ref._style)
    r=142
    for n in a.get("leaseNotes",[]):
        code=n.get("code",""); val=n.get("value",""); cite=(n.get("cite") or {})
        bc=ws.cell(r,2); cp(bc,refB); bc.value=code
        cc=ws.cell(r,3); cp(cc,refC); cc.value=cat.get(code,n.get("label",""))
        ac=ws.cell(r,1); cp(ac,refA); ac.value=(cite.get("section") or "")[:20]
        lines=wrap_lines(val)
        for li,ln in enumerate(lines):
            ec=ws.cell(r+li,5); cp(ec,refE); ec.value=ln
        r += max(1,len(lines))+1

    wb.save(out_path)

    # ---- HARD VERIFICATION: no Marshalls fingerprint anywhere ----
    chk=openpyxl.load_workbook(out_path); cs=chk[chk.sheetnames[0]]
    hits=[]
    cotenant_ok = a["tenantName"].lower()  # allow legit neighbor names in co-tenancy
    for row in cs.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str):
                low=v.lower()
                for fp in FINGERPRINTS:
                    if fp in low:
                        # "marshall" is allowed only when it appears in a co-tenant list of THIS tenant
                            hits.append((c.coordinate, fp, v[:60]))
    return ws.title, hits

title, hits = gen(sys.argv[1], sys.argv[2])
print(f"generated '{title}' -> {sys.argv[2]}")
if hits:
    print(f"!!! {len(hits)} MARSHALLS RESIDUE HITS:")
    for h in hits[:30]: print("   ", h)
    sys.exit(2)
print("VERIFIED CLEAN: no Marshalls fingerprints.")
